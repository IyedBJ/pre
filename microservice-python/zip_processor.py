import sys
import os
import zipfile
import shutil
import json
import tempfile
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from ai_extractor import extract_financial_data
    from pdf_utils import extract_text_pdf, is_pdf_scanned, extract_text_ocr
except ImportError as e:
    print(json.dumps({"error": f"Import error: {str(e)}"}))
    sys.exit(1)
# mars 2024 → 2024-03
FRENCH_MONTHS = {
    'janvier': '01', 'fevrier': '02', 'février': '02', 'mars': '03',
    'avril': '04', 'mai': '05', 'juin': '06', 'juillet': '07',
    'aout': '08', 'août': '08', 'septembre': '09', 'octobre': '10',
    'novembre': '11', 'decembre': '12', 'décembre': '12'
}
# Extraire une date depuis du texte (PDF, nom fichier, OCR)
def find_date_in_text(text):
    if not text: return None
    text_str = str(text).lower()
    # mois écrit en français mars 2024 février-2023 
    fr_date_pattern = r'([a-zA-Zéû]+)[^\w\d]*(\d{4})'
    for match in re.finditer(fr_date_pattern, text_str):
        month_str, year = match.groups()
        if month_str in FRENCH_MONTHS and 1990 <= int(year) <= 2100:
            return f"{year}-{FRENCH_MONTHS[month_str]}"
    # date au format ISO 2024-03-15 ou 2024/03/15
    iso_match = re.search(r'(\d{4})[-/](\d{2})[-/]\d{2}', text_str)
    if iso_match:
        year, month = iso_match.group(1), iso_match.group(2)
        if 1 <= int(month) <= 12 and 1990 <= int(year) <= 2100:
            return f"{year}-{month}"
    # date au format 15-03-2024 ou 15/03/2024 ou 15_03_2024
    date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
    matches = re.findall(date_pattern, text_str)
    for m in matches:
        m1, m2, year = int(m[0]), int(m[1]), m[2]
        if 1990 <= int(year) <= 2100:
            month = m1 if m2 > 12 else m2 
            if 1 <= month <= 12:
                return f"{year}-{month:02d}"
    return None
# resultat final YYYY-MM

# Extraire une date depuis fichiers Excel (.xlsx / .xls)
def find_date_in_excel_cells(file_path, ext):
    import datetime
    try:
        if ext == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, (datetime.datetime, datetime.date)):
                            return f"{cell.year}-{cell.month:02d}"
                        if isinstance(cell, str):
                            res = find_date_in_text(cell)
                            if res: return res
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                for rx in range(sheet.nrows):
                    for cx in range(sheet.ncols):
                        cell = sheet.cell(rx, cx)
                        if cell.ctype == 3:
                            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            return f"{dt.year}-{dt.month:02d}"
                        elif cell.ctype == 1:
                            res = find_date_in_text(cell.value)
                            if res: return res
    except: pass
    return None
# traiter tout un ZIP contenant des documents RH (PDF, Excel), extraire les données financières et regrouper par date
def process_zip(zip_path, employee_name, file_type, mode=None):
    from concurrent.futures import ThreadPoolExecutor
    # créer un dossier temporaire pour extraire le ZIP
    temp_dir = tempfile.mkdtemp()
    print(f"[Python] Processing ZIP: {zip_path}, Mode: {mode}", file=sys.stderr)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
    except Exception as e:
        print(f"[Python] ERROR ZIP: {e}", file=sys.stderr)
        return []
    # Résultat final : liste de dict {filename, file_type, date_group, data...} pour chaque document traité
    all_files_info = []
    for root, _, files in os.walk(temp_dir):
        for filename in files:
            ext = os.path.splitext(filename)[1].lower()
            if ext in ['.pdf', '.xlsx', '.xls']:
                all_files_info.append((os.path.join(root, filename), filename, ext))
    # traiter chaque fichier en parallèle pour accélérer le processus (surtout pour les PDF avec OCR)
    def process_file(info):
        f_path, f_name, f_ext = info
        try:
            print(f"[Python] Extracting {f_name}...", file=sys.stderr)
            data = extract_financial_data(f_path, f_ext, mode=mode)
            data["filename"] = f_name
            data["file_type"] = file_type
            
            det_date = data.get("date") or find_date_in_text(f_name)
            if not det_date and f_ext in ['.xlsx', '.xls']:
                det_date = find_date_in_excel_cells(f_path, f_ext)
            elif not det_date and f_ext == '.pdf':
                try:
                    text = extract_text_pdf(f_path) if not is_pdf_scanned(f_path) else extract_text_ocr(f_path)
                    det_date = find_date_in_text(text)
                except: pass
            data["date_group"] = det_date or "Unknown"
            return data
        except Exception as e:
            return {"filename": f_name, "error": str(e)}

    with ThreadPoolExecutor(max_workers=1) as executor:
        results = list(executor.map(process_file, all_files_info))

    shutil.rmtree(temp_dir, ignore_errors=True)
    return results

if __name__ == "__main__":
    try:
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Usage: zip_processor.py <zip_path> <name> [type] [mode]"}))
            sys.exit(1)
        
        path = sys.argv[1]
        name = sys.argv[2]
        f_type = sys.argv[3] if len(sys.argv) > 3 else "unknown"
        mode = sys.argv[4] if len(sys.argv) > 4 else None

        if not os.path.exists(path):
            print(json.dumps({"error": "File not found"}))
            sys.exit(1)

        res = process_zip(path, name, f_type, mode)
        print(json.dumps(res))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
