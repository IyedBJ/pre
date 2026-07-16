import sys
import os
import shutil
import json
import re

from pdf_utils import extract_text_pdf, extract_text_ocr, is_pdf_scanned
from nlp_utils import clean_text
from groq_utils import extract_groq, extract_employee_groq

def clean_number(value):
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, str):
        s = re.sub(r'[^\d,\.\s]', '', value).strip()
        if not s: return None
        s = s.replace(' ', '')
        
        if ',' in s and '.' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        elif '.' in s:
            if len(s.split('.')[-1]) == 3 and len(s) > 4:
                s = s.replace('.', '')
                
        try:
            return round(float(s), 2)
        except:
            return None
    return None

def extract_employee_excel(file_path, ext=None):
    """
    Profile-mode extraction for Excel files.
    Handles two formats:
      1. Nom/Prénom label cells (Notes de frais, Frais kilométriques)
      2. Désignation column with 'Salaire mensuel ... - Prénom Nom' pattern (Factures)
    Returns { nom: "Prénom NOM" } or { nom: None }
    """
    if not ext:
        ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == '.xlsx':
            try:
                import openpyxl
            except ImportError:
                return {"nom": None}
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheets = wb.worksheets
        elif ext == '.xls':
            try:
                import xlrd
            except ImportError:
                return {"nom": None}
            _wb = xlrd.open_workbook(file_path)
            # Convert to list-of-rows for uniform handling
            sheets = None  # handled separately below
        else:
            return {"nom": None}

        def get_rows_xlsx(wb):
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    yield [str(c).strip() if c is not None else '' for c in row]

        def get_rows_xls(wb):
            import xlrd
            for i in range(wb.nsheets):
                s = wb.sheet_by_index(i)
                for rx in range(s.nrows):
                    yield [str(s.cell_value(rx, cx)).strip() for cx in range(s.ncols)]

        if ext == '.xlsx':
            wb_loaded = openpyxl.load_workbook(file_path, data_only=True)
            rows_iter = get_rows_xlsx(wb_loaded)
        else:
            wb_loaded = xlrd.open_workbook(file_path)
            rows_iter = get_rows_xls(wb_loaded)

        nom    = None
        prenom = None
        desig_col = None

        all_rows = list(rows_iter)

        for ri, row in enumerate(all_rows):
            for ci, cell in enumerate(row):
                cell_str = str(cell).strip()
                cell_low = cell_str.lower()

                # Keywords for name/employee
                keywords = r'(?:nom|salari[eé]|collaborateur|agent|b[eé]n[eé]ficiaire|intervenant|nom\s*&\s*pr[eé]nom)\s*:?'
                if re.search(keywords, cell_low):
                    if ':' in cell_str:
                        val = cell_str.split(':', 1)[1].strip()
                        if val and len(val) > 2:
                            nom = val
                    if not nom:
                        for offset in range(1, min(6, len(row) - ci)):
                            candidate = str(row[ci + offset]).strip()
                            if candidate and candidate.lower() not in (':', '', 'none'):
                                nom = candidate
                                break
                
                if re.fullmatch(r'pr[eé]nom\s*:?', cell_low):
                    for offset in range(1, min(6, len(row) - ci)):
                        candidate = str(row[ci + offset]).strip()
                        if candidate and candidate.lower() not in (':', '', 'none'):
                            prenom = candidate
                            break
                
                if re.fullmatch(r'd[ée]signation', cell_low):
                    desig_col = ci

        if not nom and desig_col is not None:
            salary_pattern = re.compile(r'salaire\s+mensuel.*?[-–]\s*([a-zA-ZÀ-ÿ]+(?:\s+[a-zA-ZÀ-ÿ]+)*)', re.IGNORECASE)
            for row in all_rows:
                if desig_col < len(row):
                    cell = str(row[desig_col])
                    m = salary_pattern.search(cell)
                    if m:
                        nom = m.group(1).strip().title()
                        break

        if not nom and not prenom:
            fname = os.path.basename(file_path).lower()
            name_match = re.search(r'(?:_| -)([a-zà-ÿ]{3,}(?:[_\s-][a-zà-ÿ]{2,})+)', fname)
            if name_match: nom = name_match.group(1).replace('_', ' ').replace('-', ' ').title()

        if nom or prenom:
            parts = []
            if prenom: parts.append(prenom.strip().title())
            if nom:    parts.append(nom.strip().upper() if not prenom else nom.strip().title())
            return {"nom": " ".join(parts) if parts else None}

        return {"nom": None}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"nom": None}


def extract_excel_data(file_path, ext=None):
    if not ext:
        ext = os.path.splitext(file_path)[1].lower()
    
    all_numbers = []
    temp_path = None
    
    try:
        if not file_path.lower().endswith(tuple(['.xlsx', '.xls', '.xlsm', '.xltx', '.xltm'])):
            temp_path = file_path + ext
            shutil.copy(file_path, temp_path)
            target_path = temp_path
        else:
            target_path = file_path

        result_data = {"total": 0}

        if ext == '.xlsx':
            try:
                import openpyxl
            except ImportError:
                return {"error": "Module 'openpyxl' non trouvé."}
            
            wb = openpyxl.load_workbook(target_path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for col_idx, value in enumerate(row):
                        num = clean_number(value)
                        if num is not None: all_numbers.append(num)
                        
                        if isinstance(value, str) and value.strip().lower() == "date":
                            for offset in range(1, 4):
                                if col_idx + offset < len(row) and row[col_idx + offset]:
                                    val_date = row[col_idx + offset]
                                    import datetime
                                    if isinstance(val_date, (datetime.datetime, datetime.date)):
                                        result_data["date"] = f"{val_date.year}-{val_date.month:02d}"
                                    elif isinstance(val_date, str):
                                        import re
                                        date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
                                        m = re.search(date_pattern, val_date)
                                        if m:
                                            m1, m2, yr = int(m.group(1)), int(m.group(2)), m.group(3)
                                            mo = m1 if m2 > 12 else m2
                                            result_data["date"] = f"{yr}-{mo:02d}"

                        if isinstance(value, str) and "total à verser" in value.lower():
                            for offset in range(1, 4):
                                if col_idx + offset < len(row):
                                    n = clean_number(row[col_idx + offset])
                                    if n: return {"total": n}

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for col_idx, value in enumerate(row, 1):
                        if isinstance(value, str) and "total en euro" in value.lower():
                            last_found = 0
                            for r in range(row_idx + 1, sheet.max_row + 1):
                                val = sheet.cell(row=r, column=col_idx).value
                                n = clean_number(val)
                                if n: last_found = n
                            if last_found: result_data["total"] = last_found

        elif ext == '.xls':
            try:
                import xlrd
            except ImportError:
                return {"error": "Module 'xlrd' non trouvé."}
                
            wb = xlrd.open_workbook(target_path)
            for sheet in wb.sheets():
                for rx in range(sheet.nrows):
                    row = [sheet.cell_value(rx, cx) for cx in range(sheet.ncols)]
                    for cx, value in enumerate(row):
                        num = clean_number(value)
                        if num is not None: all_numbers.append(num)
                        
                        if isinstance(value, str) and value.strip().lower() == "date":
                            for offset in range(1, 4):
                                if cx + offset < len(row) and row[cx + offset]:
                                    val_date = row[cx + offset]
                                    try:
                                        if isinstance(val_date, float):
                                            dt = xlrd.xldate_as_datetime(val_date, wb.datemode)
                                            result_data["date"] = f"{dt.year}-{dt.month:02d}"
                                        elif isinstance(val_date, str):
                                            import re
                                            date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
                                            m = re.search(date_pattern, val_date)
                                            if m:
                                                m1, m2, yr = int(m.group(1)), int(m.group(2)), m.group(3)
                                                mo = m1 if m2 > 12 else m2
                                                result_data["date"] = f"{yr}-{mo:02d}"
                                    except:
                                        pass

                        if isinstance(value, str) and "total à verser" in value.lower():
                            for offset in range(1, 4):
                                if cx + offset < len(row):
                                    n = clean_number(row[cx + offset])
                                    if n: return {"total": n}

                for rx in range(sheet.nrows):
                    for cx in range(sheet.ncols):
                        value = sheet.cell_value(rx, cx)
                        if isinstance(value, str) and "total en euro" in value.lower():
                            last_found = 0
                            for r in range(rx + 1, sheet.nrows):
                                n = clean_number(sheet.cell_value(r, cx))
                                if n: last_found = n
                            if last_found: result_data["total"] = last_found
        
        if result_data.get("total") == 0 and all_numbers:
            result_data["total"] = float(max(all_numbers))
            
        return result_data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"Erreur Excel: {str(e)}"}
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

def process_text_data(text, mode=None):
    """
    Traite le texte pour extraire les données.
    Si mode='profile', n'extrait que les infos salariés (plus rapide).
    Par défaut, extrait tout (Finances + Profil si possible).
    """
    if mode == "profile":
        result = extract_employee_groq(text)
        result["raw_text"] = text[:5000] # Pour Regex frontend (fallback)
        return result
        
    text_clean = clean_text(text)
    
    # On extrait les finances par défaut
    result = extract_groq(text_clean)
    
    # On tente d'extraire AUSSI le nom pour l'auto-matching, sauf si mode spécifique
    if mode != 'financial_only':
        profile = extract_employee_groq(text)
        result["nom"] = profile.get("nom")
        result["numSécu"] = profile.get("numSécu")
        result["adresse"] = profile.get("adresse")
        result["raw_text"] = text[:5000] # Ajouté pour le matching frontend
        
    return result

def extract_financial_data(file_path, file_ext, mode=None):
    if file_ext in ['.xlsx', '.xls']:
        if mode == 'profile':
            return extract_employee_excel(file_path, file_ext)
        
        result = extract_excel_data(file_path, file_ext)
        if mode != 'financial_only':
            profile = extract_employee_excel(file_path, file_ext)
            result.update(profile)
        return result
    
    if is_pdf_scanned(file_path):
        text = extract_text_ocr(file_path) 
    else:
        text = extract_text_pdf(file_path)
        # Heuristique : si pdfplumber extrait très peu de texte, on tente l'OCR
        if len(text.strip()) < 100:
            import sys
            sys.stderr.write(f"[AI Extractor] Texte trop court ({len(text)} car.), tentative OCR de secours...\n")
            text = extract_text_ocr(file_path)

    import sys
    sys.stderr.write(f"[AI Extractor] Texte extrait : {len(text)} caractères\n")
    return process_text_data(text, mode)

if __name__ == "__main__":
    import os
    file_path = sys.argv[1]
    file_ext = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(file_path)[1].lower()
    mode = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = extract_financial_data(file_path, file_ext, mode)
    
    print(json.dumps(result))